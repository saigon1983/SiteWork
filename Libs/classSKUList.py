# Класс SKUList представляет из себя список любого количества SKU. Для работы с классом предоставляются многие методы из
# класса list, а так же некоторые свои специфические

import openpyxl, os, pickle
from Libs.classSKU import *



class SKUList:
    IGNORE_HEADERS = ('ТОВАРНАЯ ГРУППА 3','ТОВАРНАЯ ГРУППА 4','ТОВАРНАЯ ГРУППА 5','НАЗВАНИЕ','АРТИКУЛ','ТИП ТОВАРА',
                      'КОЭФ.',"ЕД. ИЗМЕРЕНИЯ", "ЗАРЕЗЕРВИРОВАНО", "ДАТА ИЗМ.", "АКТИВНОСТЬ", "СОРТ.", "ID",
                      "УМЕНЬШАТЬ КОЛИЧЕСТВО ПРИ ЗАКАЗЕ", "ДОСТУПНОЕ КОЛИЧЕСТВО", "ДОСТУПНОСТЬ", "КОЛИЧЕСТВО ПОДПИСОК",
                      "НЕТ СИЦ", "ЦВЕТ")

    def __init__(self, someList = None):
        # Конструктор принимает в качестве аргумента список или кортеж. Если аргумент явно не передается, создается пустой
        # экземпляр, который в последствии может быть заполнен методом append
        if type(someList) == tuple:     self.array = list(someList)
        elif type(someList) == list:    self.array = someList[:]
        elif not someList:              self.array = []
        else:                           raise TypeError ('Wrong data passed to SKUList constructor!')
        self.counter = 0
        self.setupArticlesArray()
        self.setType()
# ========== Перегруженные методы ==========
    def __str__(self):
        # Метод строкового представления объекта списка товаров
        result = ''
        for item in self.array:
            result += str(item)
            result += '\n'
        return result
    def __len__(self):
        # Перегружаем метод определения длины списка
        return len(self.array)
    def __iter__(self):
        # Метод для реализации итерации списка SKU в цикле for
        return self
    def __next__(self):
        # Метод, определяющий следующий элемент итерации
        if self.counter < len(self.array):
            self.counter += 1
            return self.array[self.counter - 1]
        else:
            self.counter = 0
            raise StopIteration
# ========== Методы настройки объекта ==========
    def setupArticlesArray(self):
        # Метод создает отдельный список артикулов списка SKU для более быстрого поиска элементов по артикулу
        self.articlesArray = []
        for sku in self.array:
            self.articlesArray.append(sku.article)
    def setType(self):
        # Метод утановки типа списка. Для гомогенных списков тип равен наименованию ТГ2, для остальных - Смешанный
        if self.isHomogenous():
            tg1type = self.array[0].TG['ТГ1']
            tg2type = self.array[0].TG['ТГ2']
            if tg2type == 'Подогреватели, вакууматоры' or self.types[1] == 'Панели домино':
                print()
                tg2type = ' '.join([X.capitalize() for X in tg2type.split()])
            if tg1type == 'Встраиваемая техника':
                tg2type += ' ВСТР'
            self.types = (tg1type, tg2type)
        else:
            self.types = ('Неопределена', 'Смешанный')
# ========== Методы изменения объекта ==========
    def append(self, item):
        # Метод добавления элемента в список. Элемент обязан быть либо экземпляром класса SKU, либо списком или кортежем
        # экземпляров этого класса. Так же item может быть другим объектом SKUList
        if type(item) == SKU:
            self.array.append(item)
            self.articlesArray.append(item.article)
            return
        elif type(item) == SKUList:
            data = item.array
        elif type(item) == list:
            data = item
        elif type(item) == tuple:
            data = list(item)
        else:
            raise TypeError ('Wrong data passed to append() method of SKUList object!')
        self.array += data
        for item in data:
            self.articlesArray.append(item.article)
        self.setType()
    def remove(self, item):
        # Метод удаляет указанный элемент из списка
        try:
            index = self.array.index(item)
            self.array.remove(item)
            self.articlesArray.remove(self.articlesArray[index])
        except:
            raise ValueError ('No such item in SKUList object!')
        self.setType()
    def removeAllByArticle(self, article):
        # Метод удаляет ВСЕ элементы списка, идетничные указанному элементу item
        try:
            finds = self.articlesArray.count(article)
            for i in range(finds):
                index = self.articlesArray.index(article)
                self.articlesArray.remove(article)
                self.array.remove(self.array[index])
        except:
            raise ValueError ('No SKU with article {} in SKUList object!'.format(article))
        self.setType()
# ========== Методы получения элементов и выборок ==========
    def getIndexByItem(self, item):
        # Метод возвращает индекс указанного SKU
        try:
            return self.array.index(item)
        except:
            raise ValueError ('No such item in SKUList object!')
    def getIndexByArticle(self, article):
        # Метод возвращает индекс SKU по ее артикулу
        if type(article) != str: raise TypeError ('Article, passed to SKUList.getIndexByArticle() must be str, not {}!'.format(type(article)))
        try:
            return self.articlesArray.index(article)
        except:
            raise ValueError ('No such item in SKUList object!')
    def getItemByIndex(self, index):
        # Метод возвращает объект SKU по указанному индексу
        if type(index) != int: raise TypeError ('Index, passed to SKUList.getItemByIndex() must be int, not {}!'.format(type(index)))
        try:
            return self.array[index]
        except:
            raise ValueError ('Wrong index argument, passed to SKUList.getItemByIndex()')
    def getItemByArticle(self, article):
        # Метод возвращает SKU по указанному артикулу, если таковая существует, или None в ином случае
        if type(article) != str: raise TypeError ('Article, passed to SKUList.getItemByArticle() must be str, not {}!'.format(type(article)))
        try:
            return self.array[self.articlesArray.index(article)]
        except:
            return None
    def getItemsByColor(self, color):
        # Метод возвращает список SKU, имеющих цвет color
        colorArray = [item for item in self.array if item.color.upper() == color.upper()]
        return SKUList(colorArray)
    def getItemsByBrand(self, brand):
        # Метод возвращает список SKU, имеющих бренд brand
        brandArray = [item for item in self.array if item.brand.upper() == brand.upper()]
        return SKUList(brandArray)
    def getItemsByGroup(self, groupName = "", groupType = None):
        # Метод возвращает спикок SKU, имеющих указанную товарную группу. В атрибуте groupName указывается наименование
        # товарной группы, в атрибуте groupType указывается порядковый номер ТГ (от 1 до 5)
        resultArray = []
        if groupName and not groupType:
            for item in self.array:
                groups = []
                for value in item.TG.values():
                    if value: groups.append(value.upper())
                if groupName.upper() in groups:
                    resultArray.append(item)
        if groupType and not groupName:
            for item in self.array:
                if item.TG['ТГ{}'.format(groupType)]: resultArray.append(item)
        if groupType and groupName:
            for item in self.array:
                if item.TG['ТГ{}'.format(groupType)].upper() == groupName.upper():
                    resultArray.append(item)
                    print(item.TG['ТГ{}'.format(groupType)].upper())
        return SKUList(resultArray)
    def getItemsByParameters(self, key, value):
        # Метод возвращает список SKU позиций с заданным свойством
        resultArray = []
        for item in self.array:
            try:
                if item.parameters[key][value]: resultArray.append(item)
            except:
                continue
        return SKUList(resultArray)
    def getIntersectsWith(self, other):
        # Метод возвращает список SKU, которые присутствуют в списке other
        resultArray = []
        for item in self.array:
            if item in other: resultArray.append(item)
        return SKUList(resultArray)
    def getExclusionsWith(self, other):
        # Метод возвращает список SKU, которые НЕ присутствуют в списке other
        resultArray = []
        for item in self.array:
            if item not in other: resultArray.append(item)
        return SKUList(resultArray)
    def splitToHomogenous(self):
        # Метод разделяет текущий список на несколько гомогенных (по ТГ2) списков и возвращает словарь списков
        newArray = {}
        newArray['Прочие'] = SKUList()
        for item in self.array:
            tg2 = item.TG['ТГ2']
            if tg2:
                if tg2 not in newArray.keys():
                    newArray[tg2] = SKUList()
                newArray[tg2].append(item)
            else:
                newArray['Прочие'].append(item)
        for someList in newArray.values():
            someList.setType()
        return newArray
# ========== Методы установки значений SKU ==========
    def setTG1(self, TG):
        # Метод устанавливает все элементам набора ТГ1 равную TG
        for item in self.array:
            item.TG['ТГ1'] = TG
    def setTG2(self, TG):
        # Метод устанавливает все элементам набора ТГ2 равную TG
        for item in self.array:
            item.TG['ТГ2'] = TG
    def setTG3(self, TG):
        # Метод устанавливает все элементам набора ТГ3 равную TG
        for item in self.array:
            item.TG['ТГ3'] = TG
    def setTG4(self, TG):
        # Метод устанавливает все элементам набора ТГ4 равную TG
        for item in self.array:
            item.TG['ТГ4'] = TG
    def setTG5(self, TG):
        # Метод устанавливает все элементам набора ТГ5 равную TG
        for item in self.array:
            item.TG['ТГ5'] = TG
# ========== Методы сохранения набора ==========
    def isHomogenous(self):
        # Метод возвращает True, если все его SKU принадлежат одной ТГ2
        if not self.array: return False
        tg1 = self.array[0].TG['ТГ1']
        tg2 = self.array[0].TG['ТГ2']
        for item in self.array:
            if item.TG['ТГ2']   != tg2: return False
            elif item.TG['ТГ1'] != tg1: return False
        return True
    def saveToExcel(self, filename):
        HEADERS = ['Артикул','Тип прибора','Бренд','Модель','ТГ1','ТГ2','ТГ3','ТГ4','ТГ5','Цвет']
        workBook = openpyxl.Workbook()
        workSheet = workBook.active
        for i in range(len(HEADERS)):
            workSheet['{}1'.format(openpyxl.utils.get_column_letter(i+1))] = HEADERS[i]
        for i in range(len(self.array)):
            workSheet['A{}'.format(i+2)] = self.articlesArray[i]
            workSheet['B{}'.format(i+2)] = self.array[i].type
            workSheet['C{}'.format(i+2)] = self.array[i].brand
            workSheet['D{}'.format(i+2)] = self.array[i].model
            workSheet['E{}'.format(i+2)] = self.array[i].TG['ТГ1']
            workSheet['F{}'.format(i+2)] = self.array[i].TG['ТГ2']
            workSheet['G{}'.format(i+2)] = self.array[i].TG['ТГ3']
            workSheet['H{}'.format(i+2)] = self.array[i].TG['ТГ4']
            workSheet['I{}'.format(i+2)] = self.array[i].TG['ТГ5']
            workSheet['J{}'.format(i+2)] = self.array[i].color
        if self.isHomogenous() and self.types[1] != 'Смешанный':
            PARASECTIONS = ('ОПИСАНИЕ ПРИБОРА','ОПИСАНИЕ','ОСНОВНЫЕ ХАРАКТЕРИСТИКИ','БЕЗОПАСНОСТЬ','ФУНКЦИОНАЛ', 'МОРОЗИЛЬНАЯ КАМЕРА')
            groupsFile = ConfigObj('DataFolder\\groupsParameters.ini')
            i = 0
            for fieldGroup in PARASECTIONS:
                if fieldGroup in groupsFile[self.types[1]]:
                    fieldName = fieldGroup
                    params = groupsFile[self.types[1]][fieldName]
                    if type(params) == str:
                        workSheet['{}1'.format(openpyxl.utils.get_column_letter(i+len(HEADERS)+1))] = params
                        i += 1
                    else:
                        for name in groupsFile[self.types[1]][fieldName]:
                            workSheet['{}1'.format(openpyxl.utils.get_column_letter(i+len(HEADERS)+1))] = name
                            i += 1
        workBook.save('Output\\' + filename + '.xlsx')
# ========== Методы класса ==========
    @classmethod
    def fromExcel(cls, excelFile, tg1, tg2):
        # Метод класса, создающий список SKU на основе предоставленной таблицы excel
        try:    workBook = openpyxl.load_workbook(excelFile)
        except: raise ValueError ('No Excel file {} found!'.format(excelFile))
        BRANDS = getBrands()
        workSheet   = workBook.get_active_sheet()
        firstRow    = workSheet[1]
        headers = {}
        for header in firstRow:
            headers[header.value] = firstRow.index(header)
        elementsList = []
        for row in workSheet.iter_rows(min_row = 2):
            skuData = {}
            skuData['Товарные группы']  = {}
            skuData['Параметры']        = OrderedDict()
            skuData['Товарные группы']['ТГ1'] = tg1
            skuData['Товарные группы']['ТГ2'] = tg2
            if row[headers['Товарная Группа 3']].value: skuData['Товарные группы']['ТГ3'] = row[headers['Товарная Группа 3']].value.strip().upper()
            if row[headers['Товарная Группа 4']].value: skuData['Товарные группы']['ТГ4'] = row[headers['Товарная Группа 4']].value.strip().upper()
            if row[headers['Товарная Группа 5']].value: skuData['Товарные группы']['ТГ5'] = row[headers['Товарная Группа 5']].value.strip().upper()
            skuData['Артикул'] = row[headers['Артикул']].value.strip().upper()
            skuData['Название'] = row[headers['Название']].value.upper()
            if row[headers['Цвет']].value: skuData['Цвет'] = row[headers['Цвет']].value.strip().upper()
            for header in headers.keys():
                if header.upper() not in cls.IGNORE_HEADERS:
                    value = row[headers[header]].value
                    if value:
                        if type(value) == str: value = value.strip()
                    skuData['Параметры'][header] = value
            elementsList.append(SKU.fromData(skuData))
        return cls(elementsList)

class Database(SKUList):
    # Класс Database представляет из себя обощенную базу всех известных SKU с некоторыми дополнительными методами доступа
    def __init__(self):
        super().__init__()
    def saveDatabase(self):
        with open('DataFolder\\database.db', 'wb') as databaseFile:
            pickle.dump(self, databaseFile)
    @classmethod
    def loadDatabase(cls):
        with open('DataFolder\\database.db', 'rb') as databaseFile:
            DB = pickle.load(databaseFile)
        return DB
    @classmethod
    def construct(cls):
        newDatabase = Database()
        for dirlist in os.walk('DataFolder\\ExcelFiles\\Product groups'):
            for file in dirlist[-1]:
                folderPath = dirlist[0]
                tg1 = folderPath.split('\\')[-1]
                tg2 = file[:-5]
                newList = SKUList.fromExcel(os.path.join(folderPath,file), tg1, tg2)
                newDatabase.append(newList)
        return newDatabase